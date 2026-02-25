"""
Exportador de transacciones con emparejamiento de transferencias internas.
Genera ficheros Excel con toda la información.
"""
import sqlite3
from pathlib import Path
from datetime import datetime, timedelta
from typing import Optional, Dict, List, Tuple
import pandas as pd

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from src.logger import get_logger


class TransferencesMatcher:
    """Empareja transferencias internas (entrada/salida) por importe y fecha."""
    
    def __init__(self, db_path: str):
        self.db_path = db_path
        self.logger = get_logger()
    
    def calculate_counterpart_id(self, transacciones: List[Dict]) -> Dict[int, Optional[int]]:
        """
        Calcula el id_contrapartida para cada transacción.
        
        Para TRANSFERENCIA/Interna:
        - Si importe < 0, busca +X en ±3 días
        - Si importe > 0, busca -X en ±3 días
        - Emparejamiento mutuo
        
        Args:
            transacciones: Lista de dicts con transacciones
        
        Returns:
            Dict {id: id_contrapartida o None}
        """
        pairings = {}  # {id: counterpart_id}
        unmatched = []
        
        # Filtrar solo transferencias internas
        internas = [t for t in transacciones if t['tipo'] == 'TRANSFERENCIA' and t['cat1'] == 'Interna']
        
        for tx in internas:
            if tx['id'] in pairings:
                continue  # Ya emparejada
            
            tx_id = tx['id']
            fecha_tx = datetime.fromisoformat(tx['fecha'])
            importe_tx = tx['importe']
            
            # Buscar la contrapartida
            contrapart = None
            
            for candidate in internas:
                if candidate['id'] == tx_id or candidate['id'] in pairings:
                    continue
                
                cand_id = candidate['id']
                fecha_cand = datetime.fromisoformat(candidate['fecha'])
                importe_cand = candidate['importe']
                
                # Condiciones de match
                # 1. Importes opuestos
                if abs(importe_tx + importe_cand) > 0.01:  # Tolerancia 1 céntimo
                    continue
                
                # 2. Fechas dentro de ±3 días
                diferencia_dias = abs((fecha_tx - fecha_cand).days)
                if diferencia_dias > 3:
                    continue
                
                # Match encontrado
                contrapart = cand_id
                break
            
            if contrapart:
                pairings[tx_id] = contrapart
                pairings[contrapart] = tx_id  # Emparejamiento mutuo
            else:
                pairings[tx_id] = None
                unmatched.append({
                    'id': tx_id,
                    'fecha': tx['fecha'],
                    'importe': importe_tx,
                    'descripcion': tx['descripcion']
                })
        
        # Loguear transferencias sin pareja
        if unmatched:
            self.logger.add_stat('transferencias_sin_pareja', unmatched)
            self.logger.debug(f"Transferencias sin pareja encontradas: {len(unmatched)}")
        
        return pairings
    
    def get_all_transactions_with_counterpart(self) -> Tuple[List[Dict], Dict]:
        """
        Obtiene todas las transacciones de BD con id_contrapartida calculado.
        
        Returns:
            (lista de transacciones dict, dict de pairings {id: counterpart_id})
        """
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        try:
            cursor.execute('''
                SELECT id, fecha, importe, descripcion, banco, cuenta, 
                       tipo, cat1, cat2, hash
                FROM transacciones
                ORDER BY fecha ASC
            ''')
            
            transacciones = [dict(row) for row in cursor.fetchall()]
            pairings = self.calculate_counterpart_id(transacciones)
            
            # Añadir id_contrapartida a cada transacción
            for tx in transacciones:
                tx['id_contrapartida'] = pairings.get(tx['id'])
            
            self.logger.debug(f"Transacciones cargadas: {len(transacciones)}")
            return transacciones, pairings
            
        finally:
            conn.close()


class ExcelExporter:
    """Exporta transacciones a fichero Excel formateado."""
    
    def __init__(self, db_path: str):
        self.db_path = db_path
        self.logger = get_logger()
        
        if not OPENPYXL_AVAILABLE:
            self.logger.error("openpyxl no está instalado. Instala con: pip install openpyxl")
            raise ImportError("openpyxl required")
    
    def export_to_excel(self, output_dir: str = 'output') -> Path:
        """
        Exporta todas las transacciones a Excel con id_contrapartida.
        
        Args:
            output_dir: Directorio de salida
        
        Returns:
            Path del archivo generado
        """
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_file = output_path / f'transacciones_{timestamp}.xlsx'
        
        self.logger.debug(f"Iniciando exportación a Excel: {excel_file}")
        
        # Obtener transacciones con emparejamiento
        matcher = TransferencesMatcher(self.db_path)
        transacciones, _ = matcher.get_all_transactions_with_counterpart()
        
        # Crear DataFrame
        df = pd.DataFrame(transacciones)
        
        # Reordenar columnas
        columnas = ['id', 'fecha', 'importe', 'descripcion', 'banco', 'cuenta', 
                    'tipo', 'cat1', 'cat2', 'hash', 'id_contrapartida']
        df = df[columnas]
        
        # Convertir fecha a datetime para que Excel la reconozca como fecha
        df['fecha'] = pd.to_datetime(df['fecha'])
        
        # Crear fichero Excel con formato
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Transacciones', index=False)
            
            # Obtener workbook y worksheet para aplicar estilos
            workbook = writer.book
            worksheet = writer.sheets['Transacciones']
            
            # Estilos
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            center_align = Alignment(horizontal='center', vertical='center')
            
            # Aplicar estilos a header
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
            
            # Ajustar anchos de columna
            col_widths = {
                'A': 6,   # id
                'B': 12,  # fecha
                'C': 12,  # importe
                'D': 30,  # descripcion
                'E': 15,  # banco
                'F': 20,  # cuenta
                'G': 15,  # tipo
                'H': 18,  # cat1
                'I': 18,  # cat2
                'J': 15,  # hash
                'K': 6,   # id_contrapartida
            }
            
            for col, width in col_widths.items():
                worksheet.column_dimensions[col].width = width
            
            # Formatear columna fecha como DD/MM/YYYY (Excel ahora la reconoce como fecha)
            for row_idx in range(2, len(transacciones) + 2):
                cell = worksheet[f'B{row_idx}']
                cell.number_format = 'DD/MM/YYYY'
            
            # Formatear celdas de importe
            for row_idx, tx in enumerate(transacciones, start=2):
                cell = worksheet[f'C{row_idx}']
                cell.number_format = '#,##0.00'
            
            # Formatear contrapartida como número (no como texto)
            for row_idx in range(2, len(transacciones) + 2):
                cell = worksheet[f'K{row_idx}']
                if cell.value and cell.value != 'None':
                    cell.value = int(cell.value) if isinstance(cell.value, float) else cell.value
                    cell.alignment = center_align
        
        self.logger.info(f"Excel exportado: {excel_file}")
        self.logger.set_stat('output_excel', str(excel_file))
        
        return excel_file
    
    def export_to_csv(self, output_dir: str = 'output') -> Path:
        """
        Exporta transacciones a CSV con id_contrapartida.
        
        Args:
            output_dir: Directorio de salida
        
        Returns:
            Path del archivo generado
        """
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        csv_file = output_path / f'transacciones_{timestamp}.csv'
        
        self.logger.debug(f"Iniciando exportación a CSV: {csv_file}")
        
        # Obtener transacciones con emparejamiento
        matcher = TransferencesMatcher(self.db_path)
        transacciones, _ = matcher.get_all_transactions_with_counterpart()
        
        # Crear DataFrame
        df = pd.DataFrame(transacciones)
        
        # Reordenar columnas
        columnas = ['id', 'fecha', 'importe', 'descripcion', 'banco', 'cuenta', 
                    'tipo', 'cat1', 'cat2', 'hash', 'id_contrapartida']
        df = df[columnas]
        
        # Guardar CSV
        df.to_csv(csv_file, index=False, encoding='utf-8')
        
        self.logger.info(f"CSV exportado: {csv_file}")
        
        return csv_file


def cleanup_old_exports(days: int = 30, output_dir: str = 'output', logs_dir: str = 'logs'):
    """
    Borra ficheros de Excel y logs con más de 'days' días.
    
    Args:
        days: Días de retención
        output_dir: Directorio de salida
        logs_dir: Directorio de logs
    """
    logger = get_logger()
    cutoff = datetime.now() - timedelta(days=days)
    
    # Limpiar Excels
    output_path = Path(output_dir)
    if output_path.exists():
        for file in output_path.glob('transacciones_*.xlsx'):
            if datetime.fromtimestamp(file.stat().st_mtime) < cutoff:
                file.unlink()
                logger.debug(f"Eliminado Excel: {file.name}")
    
    # Limpiar logs
    logs_path = Path(logs_dir)
    if logs_path.exists():
        for file in logs_path.glob('finsense_*.log'):
            if datetime.fromtimestamp(file.stat().st_mtime) < cutoff:
                file.unlink()
                logger.debug(f"Eliminado log: {file.name}")
